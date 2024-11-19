# Fisher Brown
# Lab Section: 10
# 19 Nov 24
# Jayden Robison

import openpyxl
from openpyxl.styles import Color, PatternFill
import string
wb = openpyxl.Workbook()
sheet = wb.active
silver = Color(rgb = 'c0c0c0')
fill_s = PatternFill(patternType = 'solid',fgColor = silver)
orange = Color(rgb = 'FFA500')
fill_o = PatternFill(patternType = 'solid',fgColor = orange)
blue = Color(rgb = '0000FF')
fill_b = PatternFill(patternType = 'solid',fgColor = blue)
red = Color(rgb = 'FF474C')
fill_r = PatternFill(patternType = 'solid',fgColor = red)
black = Color(rgb = '000000')
fill_bl = PatternFill(patternType = 'solid',fgColor = black)

cells_bl = ['B11','B12','B13','C10','D9','E8','F8','G7','H7','J7','K7','L7','M7','N7','07','P7','Q7','R7','S7','T7'
            ,'U7','V7','W7','X7','Y7','Z7','M6','N5','O4','P3','Q2','R2','S2','S3','S4','S5','S6','C14','D15','E15'
            ,'F15','G15','H15','I15','J15','K15','L15','M15','N15','015','P16','Q17','R18','S19','T19','U19','V19'
            ,'W19','W18','V17','U16','T15','T14','S13','S12','L12','M12','N12','012','P12','Q12','R12','M13','N14'
            ,'U14','V13','W13','X12','X11','W11','Y11','Z11','Y13','Z14','AA15','AB15','AC15','AC14','AB13','AA12'
            ,'AB12','AC12','AD11','AD10','AD9','AC8','AC7','AC6','AC5','AC4','AC3','AC2','AB3','AB4','AA5','A6']
cells_s = ['AB5','AB6','AB7','AA7','G8','H8','I8','J8','K8','L8','M8','N8','O8','P8','Q8','R8','S8','T8','U8','V8'
           ,'W8','X8','Y8','Z8','AA8','AB8','E9','F9','J9','M9','P9','S9','V9','W9','X9','Y9','Z9','AA9','AB9','AC9'
           ,'D10','E10','J10','M10','P10','S10','W10','V10','X10','Y10','Z10','AA10','AB10','AC10','C11','D11','E11'
           ,'F11','G11','H11','I11','J11','K11','L11','M11','N11','O11','P11','Q11','R11','S11','T11','U11','V11'
           ,'AA11','AB11','AC11']
cells_o = ['Q3','R3','P4','Q4','R4','O5','P5','Q5','R5','N6','O6','P6','Q6','R6','N13','O13','P13','Q13','R13','O14'
           ,'P14','Q14','R14','S14','P15','Q15','R15','S15','Q16','R16','S16','T16','R17','S17','T17','U17','S18','T18'
           ,'U18','V18','Y12','Z12','Z13','AA13','AA14','AB14']
cells_r = ['G9','H9','I9','F10','G10','H10','I10','K9','K10','L9','L10','N9','N10','O9','O10','Q9','Q10','R9','R10','T9'
           ,'T10','U9','U10']
cells_b = ['C12','D12','E12','F12','G12','H12','I12','J12','K12','C13','D13','E13','F13','G13','H13','I13','J13','K13','L13'
           ,'D14','E14','F14','G14','H14','I14','J14','K14','L14','M14']

for chr in string.ascii_uppercase[:30]:
    sheet.column_dimensions[chr].width = 5
    for i in range(1,20):
        sheet.row_dimensions[i].height = 20
        coord = chr + str(i)
        if coord in cells_bl:
            sheet[coord].fill = fill_bl
        elif coord in cells_s:
            sheet[coord].fill = fill_s
        elif coord in cells_o:
            sheet[coord].fill = fill_o
        elif coord in cells_r:
            sheet[coord].fill = fill_r
        elif coord in cells_b:
            sheet[coord].fill = fill_b

wb.save('art.xlsx')

#numRows = 20
#numCols = 20
#baseWidth = 2.5
#for col in range(1, numCols + 1):
#    colLetter = string.ascii_uppercase[col-1]
#    sheet.column_dimensions[col_letter].width = base_width
#for row in range(1, numRows + 1):
#    sheet.row_dimensions[row].height = base_width * 6

